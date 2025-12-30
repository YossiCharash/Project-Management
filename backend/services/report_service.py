from sqlalchemy import func, select, and_
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import date
from typing import List, Dict, Any
from dateutil.relativedelta import relativedelta

from backend.models.transaction import Transaction
from backend.models.category import Category
from backend.models.project import Project
from backend.models.budget import Budget
from backend.services.budget_service import BudgetService
from backend.services.fund_service import FundService
from backend.services.project_service import calculate_start_date, calculate_monthly_income_amount
import io
import zipfile
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ReportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def project_profitability(self, project_id: int) -> dict:
        income_q = select(func.coalesce(func.sum(Transaction.amount), 0)).where(
            and_(
                Transaction.project_id == project_id, 
                Transaction.type == "Income",
                Transaction.from_fund == False  # Exclude fund transactions
            )
        )
        expense_q = select(func.coalesce(func.sum(Transaction.amount), 0)).where(
            and_(
                Transaction.project_id == project_id, 
                Transaction.type == "Expense",
                Transaction.from_fund == False  # Exclude fund transactions
            )
        )
        income_val = (await self.db.execute(income_q)).scalar_one()
        expense_val = (await self.db.execute(expense_q)).scalar_one()

        proj = (await self.db.execute(select(Project).where(Project.id == project_id))).scalar_one()

        income = float(income_val)
        expenses = float(expense_val)
        profit = income - expenses

        # Check if project has budgets or funds
        has_budget = proj.budget_monthly > 0 or proj.budget_annual > 0
        # Check fund exists
        fund_service = FundService(self.db)
        fund = await fund_service.get_fund_by_project(project_id)
        has_fund = fund is not None

        return {
            "project_id": project_id,
            "income": income,
            "expenses": expenses,
            "profit": profit,
            "budget_monthly": float(proj.budget_monthly or 0),
            "budget_annual": float(proj.budget_annual or 0),
            "has_budget": has_budget,
            "has_fund": has_fund
        }

    async def get_dashboard_snapshot(self) -> Dict[str, Any]:
        """Get comprehensive dashboard snapshot with real-time financial data"""
        # Rollback any failed transaction before starting
        try:
            await self.db.rollback()
        except Exception:
            pass  # Ignore if there's no transaction to rollback
        
        # Get all active projects
        projects_query = select(Project).where(Project.is_active == True)
        projects_result = await self.db.execute(projects_query)
        projects = list(projects_result.scalars().all())
        print(f"📋 Found {len(projects)} active projects")

        if not projects:
            return {
                "projects": [],
                "alerts": {
                    "budget_overrun": [],
                    "budget_warning": [],
                    "missing_proof": [],
                    "unpaid_recurring": [],
                    "negative_fund_balance": [],
                    "category_budget_alerts": []
                },
                "summary": {
                    "total_income": 0,
                    "total_expense": 0,
                    "total_profit": 0
                },
                "expense_categories": []
            }

        # Get current date
        current_date = date.today()

        # Initialize budget service for category budget alerts
        budget_service = BudgetService(self.db)

        # Pre-load ALL project data immediately to avoid lazy loading issues
        projects_data = []
        for project in projects:
            try:
                # Extract ALL attributes immediately while session is active
                project_dict = {
                    "id": project.id,
                    "name": project.name,
                    "description": project.description,
                    "start_date": project.start_date,
                    "end_date": project.end_date,
                    "budget_monthly": project.budget_monthly,
                    "budget_annual": project.budget_annual,
                    "num_residents": project.num_residents,
                    "monthly_price_per_apartment": project.monthly_price_per_apartment,
                    "address": project.address,
                    "city": project.city,
                    "relation_project": project.relation_project,
                    "is_parent_project": project.is_parent_project,
                    "image_url": project.image_url,
                    "is_active": project.is_active,
                    "manager_id": project.manager_id,
                    "created_at": project.created_at
                }
                projects_data.append(project_dict)
            except Exception as e:
                print(f"⚠️ Error loading project data: {e}")
                continue

        # Initialize result collections
        fund_service = FundService(self.db)

        # Calculate financial data for each project
        projects_with_finance = []
        total_income = 0
        total_expense = 0
        budget_overrun_projects = []
        budget_warning_projects = []
        missing_proof_projects = []
        unpaid_recurring_projects = []
        negative_fund_balance_projects = []  # Projects with negative fund balance
        category_budget_alerts = []  # Store category budget alerts
        category_budget_alerts = []

        # Process each project using pre-loaded data
        for proj_data in projects_data:
            project_id = proj_data["id"]
            project_start_date = proj_data["start_date"]
            project_created_at = proj_data["created_at"]
            project_budget_monthly = proj_data["budget_monthly"]
            project_budget_annual = proj_data["budget_annual"]

            # Calculate start date
            if project_start_date:
                calculation_start_date = project_start_date
            else:
                calculation_start_date = current_date - relativedelta(years=1)

            # Initialize financial variables
            yearly_income = 0.0
            yearly_expense = 0.0

            try:
                # Get income transactions
                yearly_income_query = select(func.coalesce(func.sum(Transaction.amount), 0)).where(
                    and_(
                        Transaction.project_id == project_id,
                        Transaction.type == "Income",
                        Transaction.tx_date >= calculation_start_date,
                        Transaction.tx_date <= current_date,
                        Transaction.from_fund == False
                    )
                )
                yearly_income = float((await self.db.execute(yearly_income_query)).scalar_one())
            except Exception as e:
                print(f"⚠️ Error getting income for project {project_id}: {e}")
                try:
                    await self.db.rollback()
                except Exception:
                    pass
                yearly_income = 0.0

            try:
                # Get expense transactions
                yearly_expense_query = select(func.coalesce(func.sum(Transaction.amount), 0)).where(
                    and_(
                        Transaction.project_id == project_id,
                        Transaction.type == "Expense",
                        Transaction.tx_date >= calculation_start_date,
                        Transaction.tx_date <= current_date,
                        Transaction.from_fund == False
                    )
                )
                yearly_expense = float((await self.db.execute(yearly_expense_query)).scalar_one())
            except Exception as e:
                print(f"⚠️ Error getting expenses for project {project_id}: {e}")
                try:
                    await self.db.rollback()
                except Exception:
                    pass
                yearly_expense = 0.0

            # Budget is NOT income - only actual transactions count
            # Calculate budget separately for budget overrun warnings (not for income calculation)
            # Access budget fields directly - they should already be loaded
            try:
                budget_annual = float(project.budget_annual if project.budget_annual is not None else 0)
                budget_monthly = float(project.budget_monthly if project.budget_monthly is not None else 0)
            except (AttributeError, ValueError) as e:
                # If there's an issue accessing budget fields, use defaults
                budget_annual = 0.0
                budget_monthly = 0.0
            
            # Calculate income from the monthly budget (treated as expected monthly income)
            # Calculate from project start date (or created_at if start_date not available)
            project_income = 0.0
            monthly_income = float(project.budget_monthly or 0)
            if monthly_income > 0:
                # Use project start_date if available, otherwise use created_at date
                if project.start_date:
                    income_calculation_start = project.start_date
                elif hasattr(project, 'created_at') and project.created_at:
                    income_calculation_start = project.created_at.date() if hasattr(project.created_at, 'date') else project.created_at
                else:
                    # Fallback: use calculation_start_date (which is already 1 year ago if no start_date)
                    income_calculation_start = calculation_start_date
                project_income = calculate_monthly_income_amount(monthly_income, income_calculation_start, current_date)
                yearly_income = 0.0
            
            # Income = actual transactions + project income (from monthly budget)
            # Budget is NOT included in income
            project_total_income = yearly_income + project_income
            
            profit = project_total_income - yearly_expense
            
            # Calculate profit percentage based on total income
            if project_total_income > 0:
                profit_percent = (profit / project_total_income * 100)
            else:
                profit_percent = 0

            # Determine status color based on profit percentage
            if profit_percent >= 10:
                status_color = "green"
            elif profit_percent <= -10:
                status_color = "red"
            else:
                status_color = "yellow"

            # Check for budget overrun and warnings
            # Calculate expected budget for the period (same logic as budget_income)
            yearly_budget = 0.0
            # Prioritize monthly budget if both are set
            if budget_monthly > 0:
                # Same logic as budget_income calculation
                if project.start_date:
                    start_month = date(project.start_date.year, project.start_date.month, 1)
                else:
                    start_month = date(calculation_start_date.year, calculation_start_date.month, 1)
                end_month = date(current_date.year, current_date.month, 1)
                month_count = 0
                temp_month = start_month
                while temp_month <= end_month:
                    month_count += 1
                    if temp_month.month == 12:
                        temp_month = date(temp_month.year + 1, 1, 1)
                    else:
                        temp_month = date(temp_month.year, temp_month.month + 1, 1)
                yearly_budget = budget_monthly * month_count
            elif budget_annual > 0:
                # If only annual budget is set (and no monthly), calculate proportionally
                days_in_period = (current_date - calculation_start_date).days + 1
                days_in_year = 365
                yearly_budget = (budget_annual / days_in_year) * days_in_period
            if yearly_budget > 0:  # Only check if there's a budget
                budget_percentage = (yearly_expense / yearly_budget) * 100
                if yearly_expense > yearly_budget:
                    budget_overrun_projects.append(project.id)
                elif budget_percentage >= 70:  # Approaching budget (70% or more)
                    budget_warning_projects.append(project.id)

            # Check for missing proof (transactions without file_path, excluding fund transactions)
            missing_proof_query = select(func.count(Transaction.id)).where(
                and_(
                    Transaction.project_id == project.id,
                    Transaction.file_path.is_(None),
                    Transaction.tx_date >= calculation_start_date,
                    Transaction.tx_date <= current_date,
                    Transaction.from_fund == False  # Exclude fund transactions
                )
            )
            try:
                missing_proof_count = (await self.db.execute(missing_proof_query)).scalar_one()
                if missing_proof_count > 0:
                    missing_proof_projects.append(project.id)
            except Exception:
                # If query fails, rollback and continue
                try:
                    await self.db.rollback()
                except Exception:
                    pass

            # Check for unpaid recurring expenses (simplified - could be enhanced, excluding fund transactions)
            unpaid_recurring_query = select(func.count(Transaction.id)).where(
                and_(
                    Transaction.project_id == project.id,
                    Transaction.type == "Expense",
                    Transaction.is_exceptional == False,
                    Transaction.tx_date < current_date,
                    Transaction.file_path.is_(None),
                    Transaction.from_fund == False  # Exclude fund transactions
                )
            )
            try:
                unpaid_recurring_count = (await self.db.execute(unpaid_recurring_query)).scalar_one()
                if unpaid_recurring_count > 0:
                    unpaid_recurring_projects.append(project.id)
            except Exception:
                # If query fails, rollback and continue
                try:
                    await self.db.rollback()
                except Exception:
                    pass
            
            # Check for category budget alerts
            try:
                project_budget_alerts = await budget_service.check_category_budget_alerts(
                    project.id, 
                    current_date
                )
                category_budget_alerts.extend(project_budget_alerts)
            except Exception:
                # If budget checking fails, rollback and continue without it
                # This prevents the transaction from being in a failed state
                try:
                    await self.db.rollback()
                except Exception:
                    pass

            # Check for negative fund balance
            try:
                fund = await fund_service.get_fund_by_project(project.id)
                if fund and float(fund.current_balance) < 0:
                    negative_fund_balance_projects.append(project.id)
            except Exception:
                # If fund check fails, rollback and continue
                try:
                    await self.db.rollback()
                except Exception:
                    pass
                # Continue without fund balance check for this project

            # Build project data
            project_data = {
                "id": project_id,
                "name": proj_data["name"],
                "description": proj_data["description"],
                "start_date": proj_data["start_date"].isoformat() if proj_data["start_date"] else None,
                "end_date": proj_data["end_date"].isoformat() if proj_data["end_date"] else None,
                "budget_monthly": float(proj_data["budget_monthly"] or 0),
                "budget_annual": float(proj_data["budget_annual"] or 0),
                "num_residents": proj_data["num_residents"],
                "monthly_price_per_apartment": float(proj_data["monthly_price_per_apartment"] or 0),
                "address": proj_data["address"],
                "city": proj_data["city"],
                "relation_project": proj_data["relation_project"],
                "is_parent_project": proj_data["is_parent_project"],
                "image_url": proj_data["image_url"],
                "is_active": proj_data["is_active"],
                "manager_id": proj_data["manager_id"],
                "created_at": proj_data["created_at"].isoformat() if proj_data["created_at"] else None,
                "income_month_to_date": project_total_income,
                "expense_month_to_date": yearly_expense,
                "profit_percent": round(profit_percent, 1),
                "status_color": status_color,
                "budget_monthly": float(project.budget_monthly or 0),
                "budget_annual": float(project.budget_annual or 0),
                "children": []
            }

            projects_with_finance.append(project_data)
            total_income += project_total_income  # project_total_income includes budgets
            total_expense += yearly_expense

        # Build project hierarchy
        project_map = {p["id"]: p for p in projects_with_finance}
        root_projects = []

        for project_data in projects_with_finance:
            if project_data["relation_project"] and project_data["relation_project"] in project_map:
                parent = project_map[project_data["relation_project"]]
                parent["children"].append(project_data)
            else:
                root_projects.append(project_data)

        # Calculate total profit
        total_profit = total_income - total_expense

        # Get expense categories breakdown (from earliest project start_date or 1 year ago)
        # Calculate the earliest calculation_start_date across all projects
        earliest_start = date.today() - relativedelta(years=1)
        for project in projects:
            project_start = calculate_start_date(project.start_date)
            if project_start < earliest_start:
                earliest_start = project_start
        
        expense_categories_query = select(
            Category.name.label('category'),
            func.coalesce(func.sum(Transaction.amount), 0).label('total_amount')
        ).outerjoin(
            Category, Transaction.category_id == Category.id
        ).where(
            and_(
                Transaction.type == "Expense",
                Transaction.tx_date >= earliest_start,
                Transaction.tx_date <= current_date,
                Transaction.from_fund == False  # Exclude fund transactions
            )
        ).group_by(Category.name)
        
        expense_categories = []
        try:
            expense_categories_result = await self.db.execute(expense_categories_query)
            for row in expense_categories_result:
                if row.total_amount > 0:  # Only include categories with expenses
                    expense_categories.append({
                        "category": row.category or "אחר",
                        "amount": float(row.total_amount),
                        "color": self._get_category_color(row.category)
                    })
        except Exception:
            # If query fails, rollback and continue with empty categories
            try:
                await self.db.rollback()
            except Exception:
                pass

        return {
            "projects": projects_with_finance,  # Return all projects, not just root ones
            "alerts": {
                "budget_overrun": budget_overrun_projects,
                "budget_warning": budget_warning_projects,
                "missing_proof": missing_proof_projects,
                "unpaid_recurring": unpaid_recurring_projects,
                "negative_fund_balance": negative_fund_balance_projects,
                "category_budget_alerts": category_budget_alerts
            },
            "summary": {
                "total_income": round(total_income, 2),
                "total_expense": round(total_expense, 2),
                "total_profit": round(total_profit, 2)
            },
            "expense_categories": expense_categories
        }

    def _get_category_color(self, category: str) -> str:
        """Get color for expense category"""
        color_map = {
            "ניקיון": "#3B82F6",  # blue-500
            "חשמל": "#F59E0B",    # amber-500
            "ביטוח": "#8B5CF6",   # violet-500
            "גינון": "#059669",   # emerald-500
            "אחר": "#EF4444"      # red-500
        }
        return color_map.get(category, "#6B7280")  # gray-500 as default

    async def get_project_expense_categories(self, project_id: int) -> List[Dict[str, Any]]:
        """Get expense categories breakdown for a specific project"""
        expense_categories_query = (
            select(
                Category.name.label('category_name'),
                func.coalesce(func.sum(Transaction.amount), 0).label('total_amount')
            )
            .select_from(Transaction)
            .outerjoin(Category, Transaction.category_id == Category.id)
            .where(
                and_(
                    Transaction.project_id == project_id,
                    Transaction.type == "Expense",
                    Transaction.from_fund == False  # Exclude fund transactions
                )
            )
            .group_by(Category.name)
        )
        
        expense_categories_result = await self.db.execute(expense_categories_query)
        expense_categories = []
        
        for row in expense_categories_result:
            if row.total_amount > 0:  # Only include categories with expenses
                category_name = row.category_name or "אחר"
                expense_categories.append({
                    "category": category_name,
                    "amount": float(row.total_amount),
                    "color": self._get_category_color(category_name)
                })
        
        return expense_categories

    async def get_project_transactions(self, project_id: int) -> List[Dict[str, Any]]:
        """Get all transactions for a specific project (including recurring ones)"""
        transactions_query = select(Transaction).where(Transaction.project_id == project_id).order_by(Transaction.tx_date.desc())
        transactions_result = await self.db.execute(transactions_query)
        transactions = list(transactions_result.scalars().all())
        
        return [
            {
                "id": tx.id,
                "project_id": tx.project_id,
                "tx_date": tx.tx_date.isoformat(),
                "type": tx.type,
                "amount": float(tx.amount),
                "description": tx.description,
                "category": tx.category,
                "notes": tx.notes,
                "is_exceptional": tx.is_exceptional,
                "is_generated": getattr(tx, 'is_generated', False),
                "recurring_template_id": getattr(tx, 'recurring_template_id', None),
                "created_at": tx.created_at.isoformat() if hasattr(tx, 'created_at') and tx.created_at else None
            }
            for tx in transactions
        ]
    
    async def get_expenses_by_transaction_date(
        self,
        project_id: int | None = None,
        start_date: date | None = None,
        end_date: date | None = None
    ) -> Dict[str, Any]:
        """
        Get expenses aggregated by transaction date for dashboard.
        Shows expenses related to specific transaction dates with aggregation.
        """
        # Build query
        query = select(
            Transaction.tx_date,
            func.sum(Transaction.amount).label('total_expense'),
            func.count(Transaction.id).label('transaction_count')
        ).where(
            Transaction.type == 'Expense',
            Transaction.from_fund == False
        )
        
        # Filter by project if provided
        if project_id:
            query = query.where(Transaction.project_id == project_id)
        
        # Filter by date range if provided
        if start_date:
            query = query.where(Transaction.tx_date >= start_date)
        if end_date:
            query = query.where(Transaction.tx_date <= end_date)
        
        # Group by date and order
        query = query.group_by(Transaction.tx_date).order_by(Transaction.tx_date.desc())
        
        result = await self.db.execute(query)
        rows = result.all()
        
        # Format results
        expenses_by_date = []
        total_expense = 0.0
        total_count = 0
        
        for row in rows:
            expense_amount = float(row.total_expense)
            total_expense += expense_amount
            total_count += row.transaction_count
            
            expenses_by_date.append({
                'date': row.tx_date.isoformat(),
                'expense': expense_amount,
                'transaction_count': row.transaction_count
            })
        
        return {
            'expenses_by_date': expenses_by_date,
            'total_expense': total_expense,
            'total_transaction_count': total_count,
            'period_start': start_date.isoformat() if start_date else None,
            'period_end': end_date.isoformat() if end_date else None
        }

    async def generate_custom_report(self, options) -> bytes:
        """Generate a custom report (PDF, Excel, or ZIP) based on options"""
        # options is expected to be ReportOptions instance, but using dynamic typing to avoid circular import at module level
        from backend.schemas.report import ReportOptions
        
        # 1. Fetch data based on options
        project_id = options.project_id
        
        # Fetch basic project info
        proj = (await self.db.execute(select(Project).where(Project.id == project_id))).scalar_one()
        
        # --- Transactions ---
        transactions = []
        if options.include_transactions:
            query = select(Transaction).where(Transaction.project_id == project_id)
            if options.start_date:
                query = query.where(Transaction.tx_date >= options.start_date)
            if options.end_date:
                query = query.where(Transaction.tx_date <= options.end_date)
            if options.transaction_types:
                # Assuming options.transaction_types is a list like ['Income', 'Expense']
                query = query.where(Transaction.type.in_(options.transaction_types))
            if options.only_recurring:
                query = query.where(Transaction.recurring_template_id.isnot(None))
            
            # Filter by Categories (list of category names)
            if options.categories and len(options.categories) > 0:
                # Join with Category table to filter by name
                query = query.join(Category, Transaction.category_id == Category.id).where(Category.name.in_(options.categories))

            # Filter by Suppliers (list of supplier IDs)
            if options.suppliers and len(options.suppliers) > 0:
                query = query.where(Transaction.supplier_id.in_(options.suppliers))
            
            query = query.order_by(Transaction.tx_date.desc())
            result = await self.db.execute(query)
            transactions = list(result.scalars().all())

        # --- Budgets ---
        budgets_data = []
        if options.include_budgets:
            budget_service = BudgetService(self.db)
            budgets_data = await budget_service.get_project_budgets_with_spending(project_id, options.end_date)
            
        # --- Funds ---
        fund_data = None
        if options.include_funds:
            fund_service = FundService(self.db)
            fund_data = await fund_service.get_fund_by_project(project_id)

        # --- Summary Data ---
        summary_data = {}
        if options.include_summary:
            # Re-calculate summary based on date range if provided
            # Otherwise use the standard project_profitability (which is mostly all-time)
            # For custom report, it's better to respect the date range for income/expense
            
            # Use filters similar to transactions but for aggregation
            # ... implementation ...
            summary_data = await self.project_profitability(project_id) # Using standard for now, could be refined

        # 2. Generate Output
        if options.format == "pdf":
            return await self._generate_pdf(proj, options, transactions, budgets_data, fund_data, summary_data)
        elif options.format == "excel":
            return await self._generate_excel(proj, options, transactions, budgets_data, fund_data, summary_data)
        elif options.format == "zip":
            # For ZIP, we generate the PDF/Excel report AND include documents
            report_content = await self._generate_excel(proj, options, transactions, budgets_data, fund_data, summary_data)
            return await self._generate_zip(proj, options, report_content, transactions)
            
        raise ValueError("Invalid format")

    async def _generate_pdf(self, project, options, transactions, budgets, fund, summary) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        
        # Register Hebrew Font
        try:
            pdfmetrics.registerFont(TTFont('Hebrew', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
            font_name = 'Hebrew'
        except Exception:
            print("Warning: Hebrew font not found, using default")
            font_name = 'Helvetica'

        styles = getSampleStyleSheet()
        style_normal = ParagraphStyle('HebrewNormal', parent=styles['Normal'], fontName=font_name, encoding='utf-8', fontSize=10, alignment=1) # Center
        style_title = ParagraphStyle('HebrewTitle', parent=styles['Heading1'], fontName=font_name, encoding='utf-8', alignment=1, textColor=colors.HexColor('#1e3a8a')) # Center, Dark Blue
        style_h2 = ParagraphStyle('HebrewHeading2', parent=styles['Heading2'], fontName=font_name, encoding='utf-8', alignment=1, textColor=colors.HexColor('#1f2937')) # Center, Gray

        elements = []
        
        # Simple reversal function for basic RTL support in ReportLab
        def rev(text):
            if not text: return ""
            return text[::-1] if font_name == 'Hebrew' else text

        elements.append(Paragraph(rev(f"דוח פרויקט: {project.name}"), style_title))
        elements.append(Paragraph(rev(f"תאריך הפקה: {date.today().strftime('%d/%m/%Y')}"), style_normal))
        elements.append(Spacer(1, 20))
        
        # Summary
        if options.include_summary and summary:
            elements.append(Paragraph(rev("סיכום פיננסי"), style_h2))
            elements.append(Spacer(1, 10))
            data = [
                [rev("פירוט"), rev("סכום")],
                [rev("סה״כ הכנסות"), f"{summary['income']:,.2f} ₪"],
                [rev("סה״כ הוצאות"), f"{summary['expenses']:,.2f} ₪"],
                [rev("יתרה / רווח"), f"{summary['profit']:,.2f} ₪"],
            ]
            t = Table(data, colWidths=[200, 150])
            t.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), font_name),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#e5e7eb')), # Light gray header
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 20))

        # Fund
        if options.include_funds and fund:
            elements.append(Paragraph(rev("מצב קופה"), style_h2))
            elements.append(Spacer(1, 10))
            data = [
                [rev("יתרה נוכחית"), f"{fund.current_balance:,.2f} ₪"],
                [rev("הפקדה חודשית"), f"{fund.monthly_amount:,.2f} ₪"]
            ]
            t = Table(data, colWidths=[200, 150])
            t.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), font_name),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#dbeafe')), # Light blue
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 20))

        # Budgets
        if options.include_budgets and budgets:
            elements.append(Paragraph(rev("תקציב מול ביצוע"), style_h2))
            elements.append(Spacer(1, 10))
            budget_table_data = [[rev("קטגוריה"), rev("תקציב"), rev("נוצל"), rev("נותר")]]
            for b in budgets:
                cat_name = b['category'] if b['category'] else "כללי"
                budget_table_data.append([
                    rev(cat_name),
                    f"{b['amount']:,.2f}",
                    f"{b['spent_amount']:,.2f}",
                    f"{b['remaining_amount']:,.2f}"
                ])
            
            bt = Table(budget_table_data, colWidths=[120, 100, 100, 100])
            bt.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), font_name),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')), # Gray header
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(bt)
            elements.append(Spacer(1, 20))

        # Transactions
        if options.include_transactions and transactions:
            elements.append(Paragraph(rev("פירוט תנועות"), style_h2))
            elements.append(Spacer(1, 10))
            tx_data = [[rev("תאריך"), rev("סוג"), rev("סכום"), rev("תיאור")]]
            for tx in transactions:
                tx_type = "הכנסה" if tx.type == "Income" else "הוצאה"
                tx_desc = tx.description or ""
                # Truncate description nicely
                if len(tx_desc) > 30:
                    tx_desc = tx_desc[:27] + "..."
                
                tx_data.append([
                    str(tx.tx_date),
                    rev(tx_type),
                    f"{tx.amount:,.2f}",
                    rev(tx_desc)
                ])
            
            # Table handling for large data
            tx_table = Table(tx_data, repeatRows=1, colWidths=[80, 60, 80, 250])
            tx_table.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), font_name),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0f2fe')), # Light blue header
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('PADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(tx_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.read()

    async def _generate_excel(self, project, options, transactions, budgets, fund, summary) -> bytes:
        wb = Workbook()
        
        # remove default sheet
        wb.remove(wb.active)
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        fill_blue = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        fill_green = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        fill_purple = PatternFill(start_color="8b5cf6", end_color="8b5cf6", fill_type="solid")
        fill_orange = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
        
        # 1. Summary Sheet
        if options.include_summary and summary:
            ws = wb.create_sheet("סיכום")
            ws.sheet_view.rightToLeft = True
            
            # Title
            ws.append([f"דוח פרויקט: {project.name}"])
            ws.append([f"תאריך הפקה: {date.today().strftime('%d/%m/%Y')}"])
            ws.append([]) # Spacer

            ws.append(["פירוט", "סכום"])
            ws.append(["הכנסות", summary['income']])
            ws.append(["הוצאות", summary['expenses']])
            ws.append(["רווח / יתרה", summary['profit']])
            
            # Style headers
            for cell in ws[4]: 
                cell.font = header_font
                cell.fill = fill_blue
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-width
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            
        # 2. Fund
        if options.include_funds and fund:
            ws = wb.create_sheet("קופה")
            ws.sheet_view.rightToLeft = True
            ws.append(["יתרה נוכחית", fund.current_balance])
            ws.append(["הפקדה חודשית", fund.monthly_amount])
            
            # Style first col
            ws['A1'].font = Font(bold=True)
            ws['A2'].font = Font(bold=True)
            ws.column_dimensions['A'].width = 20

        # 3. Budgets
        if options.include_budgets and budgets:
            ws = wb.create_sheet("תקציב")
            ws.sheet_view.rightToLeft = True
            ws.append(["קטגוריה", "תקציב", "נוצל", "נותר", "סטטוס"])
            
            for cell in ws[1]: 
                cell.font = header_font
                cell.fill = fill_purple
                cell.alignment = Alignment(horizontal='center')
                
            for b in budgets:
                # category is stored as string in Budget model, so b['category'] is a string (name of category)
                cat_name = b['category'] if b['category'] else "כללי"
                status = "חריגה" if b['is_over_budget'] else "תקין"
                ws.append([
                    cat_name, 
                    b['amount'], 
                    b['spent_amount'], 
                    b['remaining_amount'],
                    status
                ])
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12

        # 4. Transactions
        if options.include_transactions and transactions:
            ws = wb.create_sheet("עסקאות")
            ws.sheet_view.rightToLeft = True
            headers = ["תאריך", "סוג", "סכום", "קטגוריה", "תיאור", "אמצעי תשלום", "הערות", "קובץ"]
            ws.append(headers)
            
            for cell in ws[1]: 
                cell.font = header_font
                cell.fill = fill_orange
                cell.alignment = Alignment(horizontal='center')
            
            for tx in transactions:
                # Need to fetch category name if lazy loaded or association proxy
                cat_name = tx.category if isinstance(tx.category, str) else (tx.category.name if tx.category else "")
                tx_type = "הכנסה" if tx.type == "Income" else "הוצאה"
                
                row = [
                    tx.tx_date, 
                    tx_type, 
                    tx.amount, 
                    cat_name,
                    tx.description or "", 
                    tx.payment_method or "", 
                    tx.notes or "",
                    "כן" if tx.file_path else "לא"
                ]
                ws.append(row)
            
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    async def _generate_zip(self, project, options, report_content, transactions) -> bytes:
        from backend.services.s3_service import S3Service
        try:
            s3_service = S3Service()
            has_s3 = True
        except Exception:
            has_s3 = False
            
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
            # Report File
            ext = "xlsx" if options.format == "zip" else "pdf" # Default to excel inside zip if zip requested directly?
            # Actually, if options.format is zip, we generated excel above.
            zf.writestr(f"report.{ext}", report_content)
            
            if has_s3:
                # Add Contract if requested
                if options.include_project_contract and project.contract_file_url:
                    try:
                        contract_content = s3_service.get_file_content(project.contract_file_url)
                        if contract_content:
                            fname = project.contract_file_url.split('/')[-1]
                            zf.writestr(f"contract_{fname}", contract_content)
                    except Exception as e:
                        print(f"Error adding contract to ZIP: {e}")

                # Add Image if requested
                if options.include_project_image and project.image_url:
                    try:
                        # Assuming image_url is a path or key relative to bucket/base
                        image_content = s3_service.get_file_content(project.image_url)
                        if image_content:
                            fname = project.image_url.split('/')[-1]
                            zf.writestr(f"project_image_{fname}", image_content)
                    except Exception as e:
                        print(f"Error adding image to ZIP: {e}")

                # Documents
                if options.include_transactions:
                    for tx in transactions:
                        if tx.file_path:
                            try:
                                content = s3_service.get_file_content(tx.file_path)
                                if content:
                                    # Safe filename
                                    fname = f"{tx.tx_date}_{tx.id}.{tx.file_path.split('.')[-1]}"
                                    zf.writestr(f"documents/{fname}", content)
                            except Exception:
                                pass

        output.seek(0)
        return output.read()

    async def generate_excel_report(self, project_id: int) -> bytes:
        """Generate Excel report for a project"""
        # Fetch data
        project_data = await self.project_profitability(project_id)
        transactions = await self.get_project_transactions(project_id)
        expense_categories = await self.get_project_expense_categories(project_id)
        
        # Get Project details
        proj = (await self.db.execute(select(Project).where(Project.id == project_id))).scalar_one()
        
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        fill_blue = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        fill_orange = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
        fill_green = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        
        # 1. Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "סיכום"
        ws_summary.sheet_view.rightToLeft = True
        
        summary_data = [
            [f"דוח פרויקט: {proj.name}"],
            [f"תאריך הפקה: {date.today().strftime('%d/%m/%Y')}"],
            [],
            ["סיכום פיננסי"],
            ["סה״כ הכנסות", project_data["income"]],
            ["סה״כ הוצאות", project_data["expenses"]],
            ["רווח", project_data["profit"]],
            ["תקציב (חודשי)", project_data["budget_monthly"]],
            ["תקציב (שנתי)", project_data["budget_annual"]],
        ]
        
        for row in summary_data:
            ws_summary.append(row)
            
        ws_summary.column_dimensions['A'].width = 20
        ws_summary['D1'].fill = fill_blue # Just an example if we had headers properly
            
        # 2. Transactions Sheet
        ws_tx = wb.create_sheet("עסקאות")
        ws_tx.sheet_view.rightToLeft = True
        headers = ["תאריך", "סוג", "סכום", "קטגוריה", "תיאור", "אמצעי תשלום", "הערות", "קובץ"]
        ws_tx.append(headers)
        for cell in ws_tx[1]:
            cell.font = header_font
            cell.fill = fill_orange
            cell.alignment = Alignment(horizontal='center')
            
        for tx in transactions:
            tx_type = "הכנסה" if tx["type"] == "Income" else "הוצאה"
            row = [
                tx["tx_date"], 
                tx_type, 
                tx["amount"], 
                tx["category"] or "", 
                tx["description"] or "", 
                tx.get("payment_method") or "", 
                tx["notes"] or "",
                "כן" if tx.get("file_path") else "לא"
            ]
            ws_tx.append(row)

        ws_tx.column_dimensions['A'].width = 12
        ws_tx.column_dimensions['B'].width = 10
        ws_tx.column_dimensions['C'].width = 12
        ws_tx.column_dimensions['D'].width = 15
        ws_tx.column_dimensions['E'].width = 30
            
        # 3. Categories Breakdown
        ws_cat = wb.create_sheet("קטגוריות")
        ws_cat.sheet_view.rightToLeft = True
        ws_cat.append(["קטגוריה", "סכום"])
        for cell in ws_cat[1]:
            cell.font = header_font
            cell.fill = fill_green
            cell.alignment = Alignment(horizontal='center')
            
        for cat in expense_categories:
            ws_cat.append([cat["category"], cat["amount"]])
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    async def generate_zip_export(self, project_id: int) -> bytes:
        """Generate ZIP export with Excel report and transaction documents"""
        # Generate Excel
        excel_data = await self.generate_excel_report(project_id)
        
        # Get transactions for documents
        transactions = await self.get_project_transactions(project_id)
        
        # Initialize S3 Service
        from backend.services.s3_service import S3Service
        try:
            s3_service = S3Service()
            has_s3 = True
        except Exception:
            has_s3 = False
            print("Warning: S3 Service not available for ZIP export")
        
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
            # Add Excel report
            zf.writestr(f"project_{project_id}_report.xlsx", excel_data)
            
            # Add documents
            if has_s3:
                for tx in transactions:
                    file_path = tx.get("file_path")
                    if file_path:
                        try:
                            # Extract filename from path or URL
                            original_filename = file_path.split("/")[-1]
                            # Use transaction ID and date to make filename unique and meaningful
                            ext = original_filename.split(".")[-1] if "." in original_filename else "bin"
                            filename = f"{tx['tx_date']}_{tx['type']}_{tx['id']}.{ext}"
                            
                            content = s3_service.get_file_content(file_path)
                            if content:
                                zf.writestr(f"documents/{filename}", content)
                        except Exception as e:
                            print(f"Error adding file {file_path} to ZIP: {e}")
                        
        output.seek(0)
        return output.read()
